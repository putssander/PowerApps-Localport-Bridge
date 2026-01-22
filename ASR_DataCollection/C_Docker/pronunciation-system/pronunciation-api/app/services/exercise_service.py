"""
Exercise Service

Manages XLSX exercise files with auto-IPA generation.
"""

import json
import logging
import re
import uuid
from pathlib import Path
from typing import Optional
from io import BytesIO

from openpyxl import load_workbook
from g2p_en import G2p

logger = logging.getLogger("pronunciation-api.exercise_service")

# ARPAbet to IPA vowel mapping
ARPABET_TO_IPA = {
    'AA': 'ɑ', 'AE': 'æ', 'AH': 'ʌ', 'AO': 'ɔ', 'AW': 'aʊ',
    'AY': 'aɪ', 'EH': 'ɛ', 'ER': 'ɜ', 'EY': 'eɪ', 'IH': 'ɪ',
    'IY': 'i', 'OW': 'oʊ', 'OY': 'ɔɪ', 'UH': 'ʊ', 'UW': 'u',
}


class ExerciseService:
    """Manages exercise XLSX files with auto-IPA phoneme generation."""
    
    def __init__(self, exercises_dir: Path):
        """
        Initialize exercise service.
        
        Args:
            exercises_dir: Directory to store exercise JSON files
        """
        self.exercises_dir = Path(exercises_dir)
        self.exercises_dir.mkdir(parents=True, exist_ok=True)
        
        # Initialize G2P for auto phoneme generation
        logger.info("Initializing G2P for auto-IPA generation...")
        self.g2p = G2p()
        logger.info("G2P ready")
    
    def _generate_phonemes(self, text: str) -> str:
        """
        Generate IPA phonemes from text using g2p-en.
        
        Converts ARPAbet output to IPA for compatibility with wav2vec2.
        
        Args:
            text: English text
            
        Returns:
            Space-separated IPA phoneme string
        """
        arpabet = self.g2p(text)
        
        # Convert to IPA
        ipa_list = []
        for p in arpabet:
            if not p.strip() or p.startswith(' '):
                continue
            # Remove stress markers (0, 1, 2)
            base_p = re.sub(r'[012]', '', p)
            if base_p in ARPABET_TO_IPA:
                ipa_list.append(ARPABET_TO_IPA[base_p])
            else:
                ipa_list.append(base_p.lower())
        
        return ' '.join(ipa_list)
    
    def import_xlsx(self, file_content: bytes, filename: str) -> dict:
        """
        Import exercise from XLSX file.
        
        Expected format:
        - Column A: Sentence text (required)
        - Column B: Expected IPA phonemes (optional)
        - Column C: Focus vowels (optional, comma-separated)
        
        Args:
            file_content: Raw XLSX file bytes
            filename: Original filename
            
        Returns:
            Dictionary with exercise_id, sentences, stats
        """
        logger.info(f"Importing exercise from: {filename}")
        
        # Load workbook
        wb = load_workbook(BytesIO(file_content))
        ws = wb.active
        
        sentences = []
        auto_generated_count = 0
        
        # Process rows (skip header if present)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            if not row or not row[0]:
                continue
            
            text = str(row[0]).strip()
            expected_phonemes = str(row[1]).strip() if len(row) > 1 and row[1] else None
            focus_vowels_str = str(row[2]).strip() if len(row) > 2 and row[2] else None
            
            # Auto-generate phonemes if not provided
            if not expected_phonemes:
                expected_phonemes = self._generate_phonemes(text)
                auto_generated_count += 1
                logger.debug(f"Auto-generated phonemes for: '{text[:30]}...'")
            
            # Parse focus vowels
            focus_vowels = None
            if focus_vowels_str:
                focus_vowels = [v.strip() for v in focus_vowels_str.split(',')]
            
            sentences.append({
                "index": row_idx,
                "text": text,
                "expected_phonemes": expected_phonemes,
                "focus_vowels": focus_vowels
            })
        
        if not sentences:
            raise ValueError("No sentences found in XLSX file")
        
        # Generate exercise ID and save
        exercise_id = str(uuid.uuid4())[:8]
        exercise_data = {
            "exercise_id": exercise_id,
            "original_filename": filename,
            "sentence_count": len(sentences),
            "auto_generated_phonemes": auto_generated_count,
            "sentences": sentences
        }
        
        # Save to JSON
        exercise_path = self.exercises_dir / f"{exercise_id}.json"
        with open(exercise_path, 'w', encoding='utf-8') as f:
            json.dump(exercise_data, f, ensure_ascii=False, indent=2)
        
        logger.info(
            f"Exercise saved: {exercise_id} | "
            f"{len(sentences)} sentences | "
            f"{auto_generated_count} auto-generated phonemes"
        )
        
        return exercise_data
    
    def list_exercises(self) -> list[dict]:
        """
        List all available exercises.
        
        Returns:
            List of exercise summaries
        """
        exercises = []
        
        for json_file in self.exercises_dir.glob("*.json"):
            try:
                with open(json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    exercises.append({
                        "exercise_id": data.get("exercise_id"),
                        "original_filename": data.get("original_filename"),
                        "sentence_count": data.get("sentence_count", 0)
                    })
            except Exception as e:
                logger.warning(f"Failed to load exercise {json_file}: {e}")
        
        return exercises
    
    def get_exercise(self, exercise_id: str) -> Optional[dict]:
        """
        Get a specific exercise by ID.
        
        Args:
            exercise_id: Exercise ID
            
        Returns:
            Exercise data or None if not found
        """
        exercise_path = self.exercises_dir / f"{exercise_id}.json"
        
        if not exercise_path.exists():
            return None
        
        with open(exercise_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    def delete_exercise(self, exercise_id: str) -> bool:
        """
        Delete an exercise.
        
        Args:
            exercise_id: Exercise ID
            
        Returns:
            True if deleted, False if not found
        """
        exercise_path = self.exercises_dir / f"{exercise_id}.json"
        
        if not exercise_path.exists():
            return False
        
        exercise_path.unlink()
        logger.info(f"Deleted exercise: {exercise_id}")
        return True
